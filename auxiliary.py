import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side

from keyboards import Keyboards

class Auxiliary:
    def __init__(self, bot, connection):
        
        self.bot = bot

        self.keyboard = Keyboards(connection)

        cursor = connection.cursor()
        self.cursor = cursor

    def extract_arg(self, arg):
        return arg.split()[1:]
    
    def filter(self, text):
        text = text.replace('_', '\_').replace('*', '\*').replace('[', '\[').replace(']', '\]').replace('(', '\(').replace(')', '\)').replace('~', '\~').replace('`', '\`').replace('>', '\>').replace('#', '\#').replace('+', '\+').replace('-', '\-').replace('=', '\=').replace('|', '\|').replace('{', '\{').replace('}', '\}').replace('.', '\.').replace('!', '\!')
        return text

    def exitStepHandler(self, message, status):
        bot = self.bot
        
        text = ""
        match status:
            case "error":
                text = "Произошла ошибка, попробуйте позже."
            case "notCompleted":
                text = "Заявка на получение пропуска *не оформлена*, так как паспортные данные не заполнены."
            case _:
                text = "👌"
        
        bot.send_message(chat_id=message.chat.id, text=text, parse_mode="Markdown", reply_markup=self.keyboard.genMainKeyboard(message.from_user.id))
        bot.clear_step_handler_by_chat_id(chat_id=message.chat.id)
        return
    
    def export_to_excel(self, query_string, headings, filepath):
        try:
            cursor = self.cursor
            
            cursor.execute(query_string)
            data = cursor.fetchall()

            wb = openpyxl.Workbook()
            sheet = wb.active

            header = NamedStyle(name="header")
            header.font = Font(bold=True)
            header.border = Border(bottom=Side(border_style="thin"))
            header.alignment = Alignment(horizontal="center", vertical="center")

            for colno, heading in enumerate(headings, start = 1):
                sheet.cell(row = 1, column = colno).value = heading

            header_row = sheet[1]
            for cell in header_row:
                cell.style = header

            for rowno, row in enumerate(data, start = 2):
                itemId = row[0]
                for colno, cell_value in enumerate(row[1:], start = 1):
                    if filepath.startswith('usersList') and colno == 17:
                        cursor.execute(f'''SELECT name FROM users_projects
                                INNER JOIN projects ON users_projects.projectid = projects.id
                                WHERE userid = '{itemId}';''')
                        projects = cursor.fetchall()
                        projectNames = ""
                        for project in projects:
                            projectNames += project[0] + ", "
                        
                        sheet.cell(row = rowno, column = colno).value = cell_value # status
                        sheet.cell(row = rowno, column = colno+1).value = projectNames[:-2] # projectsMember
                        continue
                    elif filepath.startswith('projectsList') and colno == 3:
                        cursor.execute(f'''SELECT concat(lastname, ' ', firstname, ' ', patronymic, ' (@', username, ')') as FIO FROM users_projects
                                            INNER JOIN users ON users_projects.userid = users.id
                                            WHERE projectid = {itemId} and role = 'AUTHOR';''')
                        fioAuthor = cursor.fetchone()[0]

                        cursor.execute(f'''SELECT concat(lastname, ' ', firstname, ' ', patronymic, ' (@', username, ')') FROM users_projects
                                INNER JOIN users ON users_projects.userid = users.id
                                WHERE projectid = {itemId} and role = 'PARTNER';''')
                        items = cursor.fetchall()
                        partners = ""

                        for partner in items:
                            partners += f'{partner[0] if partner[0] != None else ""};\n'
                        
                        if partners.strip() != "":
                            partners = f"{partners}"[:-2] + "."

                        sheet.cell(row = rowno, column = colno).value = cell_value
                        sheet.cell(row = rowno, column = colno+1).value = fioAuthor
                        sheet.cell(row = rowno, column = colno+2).value = partners
                        continue
                    elif filepath.startswith('vacanciesList'):
                        if colno == 1:
                            if cell_value:
                                value = "Новое"
                            else:
                                value = "Просмотрено"
                        elif colno == 2:
                            if cell_value:
                                value = "В поиске"
                            else:
                                value = "Не актуально"
                        elif colno == 7:
                            value = self.db.getProjectNameById(cell_value)
                        else: 
                            sheet.cell(row = rowno, column = colno).value = cell_value
                            continue
                        sheet.cell(row = rowno, column = colno).value = value
                        continue

                    sheet.cell(row = rowno, column = colno).value = cell_value
            wb.save(filepath)
        except Exception as e:
            print("Error in export Excel" + e)