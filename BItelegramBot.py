from telebot import types, telebot
from telebot.types import InlineKeyboardMarkup, InlineKeyboardButton
import csv
import psycopg2
import re
import config
import requests
import json
import time
from threading import Thread
import schedule
from datetime import datetime
import openpyxl
from openpyxl.styles import NamedStyle, Font, Alignment, Border, Side

bot = telebot.TeleBot(config.BOT_TOKEN)

connection = psycopg2.connect(
        host=config.HOST,
        user=config.USER,
        password=config.PASSWORD,
        database=config.DATABASE,
        port=config.PORT
    )
cursor = connection.cursor()

print('[INFO] PostgreSQL start')


# -------------------------------------------------- Update user info steps -----------------------


def process_getUsernameForUpdate_step(message):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
        if getUserById(message.from_user.id)['status'].find("ADMIN") == -1:
            bot.send_message(chat_id=message.chat.id, text= f"Недостаточно прав")
            return 
        
        username = message.text.strip('@')
        userId = getUserIdByUsernameAndFIO(username, "userName")
        if userId == "-":
            msg = bot.send_message(chat_id=message.chat.id, text= f"Пользователь не найден. Повторите ввод")
            bot.register_next_step_handler(msg, process_getUsernameForUpdate_step)
        else:
            bot.send_message(chat_id=message.chat.id, text= f"Обновление данных пользователя @{username}")
            updateFullname(message.chat.id, userId)
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_getUsernameForSelect_step(message):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
        
        userInfoMsg = message.text.strip()

        dataFormat = ""
        if userInfoMsg.startswith('@'):
            userInfoMsg = userInfoMsg.strip('@')
            dataFormat = "userName"
        elif len(userInfoMsg.split(' ')) == 1:
            dataFormat = "lastName"
        elif len(userInfoMsg.split(' ')) == 2:
            dataFormat = "lastName_firstName"
        else:
            msg = bot.send_message(chat_id=message.chat.id, text= f"Пользователь не найден. Повторите ввод")
            bot.register_next_step_handler(msg, process_getUsernameForSelect_step)
            return

        userId = getUserIdByUsernameAndFIO(userInfoMsg, dataFormat)
        if userId == "-":
            msg = bot.send_message(chat_id=message.chat.id, text= f"Пользователь не найден. Повторите ввод")
            bot.register_next_step_handler(msg, process_getUsernameForSelect_step)
            return

        if not isinstance(userId, str):        
            for user in userId:
                userInfo(message, user[0])
        else:
            userInfo(message, userId)        
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_updateFullName_step(message, userId):
    try:
        # surname = message.text.strip().split(' ')[0]
        # name = message.text.strip().replace(surname, "").strip().split(' ')[0]
        # patronymic = message.text.strip().replace(surname, "").replace(name, "").strip().split(' ')[0]

        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
            
        pattern = r'([А-ЯЁа-яё]+)\s+([А-ЯЁа-яё]+)\s+([А-ЯЁа-яё]+)'
        match = re.search(pattern, message.text.strip())
        
        if match:
            surname, name, patronymic = match.groups()
        else:
            msg = bot.reply_to(message, f'Неверный формат, повторите ввод ФИО\n(Пример: _Иванов Иван Иванович_)', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_updateFullName_step, userId)
            return 

        cursor.execute(f'''UPDATE users SET firstname='{name}',
                            lastname='{surname}', patronymic='{patronymic}'
                            WHERE id = '{userId}';''')
        connection.commit()

        msg = bot.reply_to(message, f'Укажите сферу деятельности')

        bot.register_next_step_handler(msg, process_updateFieldOfActivity_step, userId)
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_updateFieldOfActivity_step(message, userId):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return

        fieldOfActivity = message.text

        if len(fieldOfActivity) <= 1:
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод сферы деятельности\n(Пример: _Медицина и образование_)', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_updateFieldOfActivity_step, userId)
            return
        
        cursor.execute(f'''UPDATE users SET fieldofactivity='{fieldOfActivity}'
                            WHERE id = '{userId}';''')
        connection.commit()

        msg = bot.reply_to(message, 'Записано. Расскажите о себе: опишите свои области компетенций, укажите hard skills.')
        bot.register_next_step_handler(msg, process_updateAboutMe_step, userId)
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def process_updateAboutMe_step(message, userId):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
        
        aboutMe = message.text

        if len(aboutMe) <= 1:
            msg = bot.reply_to(message, 'Неверный формат, повторите')
            bot.register_next_step_handler(msg, process_updateAboutMe_step, userId)
            return
        
        cursor.execute(f'''UPDATE users SET aboutme='{aboutMe}'
                            WHERE id = '{userId}';''')
        connection.commit()

        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        reaBtn = types.KeyboardButton(text="РЭУ им. Г.В. Плеханова")
        hseBtn = types.KeyboardButton(text="НИУ ВШЭ")
        mftiBtn = types.KeyboardButton(text="МФТИ")
        goHomeBtn = types.KeyboardButton(text="↩ Выйти")
        keyboard.add(reaBtn, hseBtn, mftiBtn, goHomeBtn)

        msg = bot.send_message(message.chat.id, 'Выберите учебное заведение или укажите вручную', parse_mode="Markdown", reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_updateEducationalInstitution_step, userId) 
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def process_updateEducationalInstitution_step(message, userId):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
    
        educationalInstitution = message.text

        if len(educationalInstitution) < 3:
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод учебного заведения', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_updateEducationalInstitution_step)
            return
    
        cursor.execute(f'''UPDATE users SET educationalinstitution='{educationalInstitution}'
                            WHERE id = '{userId}';''')
        connection.commit()

        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        keyboard.add(types.KeyboardButton(text="↩ Выйти"))

        if message.text == "РЭУ им. Г.В. Плеханова":
            msg = bot.send_message(message.chat.id, 'Укажите вашу Высшую школу/Институт/Факультет', parse_mode="Markdown", reply_markup=keyboard)
            bot.register_next_step_handler(msg, process_updateFaculty_step, userId) 
            return

        cursor.execute(f'''UPDATE users SET course=NULL, 
                                faculty=NULL, direction=NULL, 
                                speciality=NULL, "group"=NULL
                                WHERE id = '{userId}';''')
        connection.commit()

        msg = bot.send_message(message.chat.id, 'Отлично, оставьте свой *номер телефона*\n(_+79993332211_)', parse_mode="Markdown", reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_updatePhoneNum_step, userId) 
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def process_updateFaculty_step(message, userId):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
    
        faculty = message.text

        if len(faculty) < 3:
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод ВШ/И/Ф', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_updateFaculty_step, userId)
            return
    
        cursor.execute(f'''UPDATE users SET faculty='{faculty}'
                            WHERE id = '{userId}';''')
        connection.commit()

        keyboard = types.ReplyKeyboardMarkup(row_width=4, resize_keyboard=True)
        keyboard.add(types.KeyboardButton(text="1"), types.KeyboardButton(text="2"), types.KeyboardButton(text="3"), types.KeyboardButton(text="4"), types.KeyboardButton(text="5"))
        keyboard.add(types.KeyboardButton(text="↩ Выйти"))

        msg = bot.send_message(message.chat.id, 'На каком *курсе* вы обучаетесь?', parse_mode="Markdown", reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_updateCourse_step, userId) 
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def process_updateCourse_step(message, userId):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
    
        course = message.text

        if len(course) > 1 or not course.isdigit():
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод курса\n(Пример: 2)', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_updateCourse_step, userId)
            return
    
        cursor.execute(f'''UPDATE users SET course={course}
                            WHERE id = '{userId}';''')
        connection.commit()

        keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
        keyboard.add(types.KeyboardButton(text="↩ Выйти"))

        msg = bot.send_message(message.chat.id, 'Укажите Ваше *направление подготовки*.', parse_mode="Markdown", reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_updateDirection_step, userId) 
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def process_updateDirection_step(message, userId):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
    
        direction = message.text

        if len(direction) < 5:
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод направления подготовки', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_updateDirection_step, userId)
            return
    
        cursor.execute(f'''UPDATE users SET direction='{direction}'
                            WHERE id = '{userId}';''')
        connection.commit()

        msg = bot.send_message(message.chat.id, 'Укажите *специальность*', parse_mode="Markdown")
        bot.register_next_step_handler(msg, process_updateSpeciality_step, userId) 
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def process_updateSpeciality_step(message, userId):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
    
        speciality = message.text

        if len(speciality) < 5:
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод специальности', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_updateSpeciality_step, userId)
            return
    
        cursor.execute(f'''UPDATE users SET speciality='{speciality}'
                            WHERE id = '{userId}';''')
        connection.commit()

        msg = bot.send_message(message.chat.id, 'Укажите *группу*, в которой Вы обучаетесь', parse_mode="Markdown")
        bot.register_next_step_handler(msg, process_updateGroup_step, userId) 
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def process_updateGroup_step(message, userId):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
    
        group = message.text

        if len(group) < 4:
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод группы', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_updateGroup_step, userId)
            return
    
        cursor.execute(f'''UPDATE users SET "group"='{group}'
                            WHERE id = '{userId}';''')
        connection.commit()

        msg = bot.send_message(message.chat.id, 'Отлично, оставьте свой номер телефона\n(_+79993332211_)', parse_mode="Markdown")
        bot.register_next_step_handler(msg, process_updatePhoneNum_step, userId) 
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def process_updatePhoneNum_step(message, userId):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return

        phoneNum = message.text

        pattern = r'^\+\d{11}$'
        if not re.match(pattern, phoneNum):
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод номера\n(Пример: _+79993332211_)', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_updatePhoneNum_step, userId)
            return
        
        cursor.execute(f'''UPDATE users SET phonenum='{phoneNum}'
                            WHERE id = '{userId}';''')
        connection.commit()

        msg = bot.reply_to(message, 'Осталось добавить почту\n(_ivanov.i.i@gmail.com_)', parse_mode="Markdown")
        bot.register_next_step_handler(msg, process_updateEmail_step, userId) 
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def process_updateEmail_step(message, userId):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
        email = message.text

        pattern = r"^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$"
        if not re.match(pattern, email):
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод почты\n(Пример: _ivanov.i.i@gmail.com_)', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_updateEmail_step, userId)
            return
        
        cursor.execute(f'''UPDATE users SET email='{email}'
                            WHERE id = '{userId}';''')
        connection.commit()


        cursor.execute(f'''SELECT name FROM users_projects
                            INNER JOIN projects ON users_projects.projectid = projects.id
                            WHERE userid = '{userId}';''')
        projects = cursor.fetchall()


        projectNames = ""
        for project in projects:
            projectNames += project[0] + ", "
        
        if projectNames.strip() != "":
            projectNames = f"*Участник проектов*: {projectNames}"[:-2]

        user = getUserById(userId)
        email = user['email'].replace("_", "\_")

        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        markup.add('🔴 Повторить ввод', '🟢 Все верно')

        additionInfo = ""
        if user['educationalinstitution'] == "РЭУ им. Г.В. Плеханова":
            additionInfo = f" ({user['course']} курс, {user['faculty']}, по направлению «{user['direction']}», специальность «{user['speciality']}», группа {user['group']})"
        bot.send_message(message.chat.id, f'''Данные обновлены, проверьте корректность:

*ФИО*: _{user['lastname'] if user['lastname'] != None else ""} {user['firstname'] if user['firstname'] != None else ""} {user['patronymic'] if user['patronymic'] != None else ""}_

*Сфера деятельности*: _{user['fieldofactivity']}_

*Учебное заведение*: _{user['educationalinstitution']}{additionInfo}_

*О себе*: _{user['aboutme']}_

*Телефон*: _{user['phonenum']}_

*Email*: {email}

{projectNames.strip()}''', 
        reply_markup=markup, 
        parse_mode="Markdown")
        
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")


# -------------------------------------------------- Create project steps -----------------------

def process_getProjectnameForSelect_step(message):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
        
        projectName = message.text.strip()
        projectId = getProjectIdByProjectname(projectName)
        if projectId == "-":
            msg = bot.send_message(chat_id=message.chat.id, text= f"Проект не найден. Повторите ввод")
            bot.register_next_step_handler(msg, process_getProjectnameForSelect_step)
            return
        projectInfo(message, projectId)
        
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_insertProjectName_step(message):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
            
        projectName = message.text
        if len(projectName) < 3:
            msg = bot.reply_to(message, f'Неверный формат, повторите ввод наименования', parse_mode="Markdown")
            bot.register_next_step_handler(message=msg, callback=process_insertProjectName_step)
            return 
        
        cursor.execute(f'''SELECT EXISTS(SELECT 1 FROM projects
                                WHERE name = '{projectName.strip()}');''')
        isAlreadyExists = cursor.fetchone()
        
        if isAlreadyExists[0]:
            msg = bot.reply_to(message, f'Проект с таким наименованием уже существет, повторите ввод', parse_mode="Markdown")
            bot.register_next_step_handler(message=msg, callback=process_insertProjectName_step)
            return 

        msg = bot.reply_to(message, f'Составьте *описание* проекта {projectName}:', parse_mode="Markdown")

        bot.register_next_step_handler(msg, process_insertProjectDescription_step, projectName)
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_insertProjectDescription_step(message, projectName):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
            
        projectDescription = message.text

        if len(projectDescription) < 3:
            msg = bot.reply_to(message, f'Неверный формат, повторите ввод описания', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_insertProjectDescription_step, projectName)
            return 

        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        goHomeBtn = types.KeyboardButton(text="↩ Выйти")
        keyboard.add(types.KeyboardButton(text='Креативные индустрии'), types.KeyboardButton(text='Медицина и образование'), types.KeyboardButton(text='Социальные'), types.KeyboardButton(text='Технологические'), types.KeyboardButton(text='Экологические'), goHomeBtn)

        msg = bot.send_message(message.from_user.id, f'Укажите *категорию* вашего проекта', parse_mode="Markdown", reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_insertProjectCategory_step, projectName, projectDescription)
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_insertProjectCategory_step(message, projectName, projectDescription):
    try:
        projectCategory = ""

        if message.text == '↩ Выйти':
            exitStepHandler(message, "ok")
            return
        
        Categories = ['Креативные индустрии', 'Медицина и образование', 'Социальные', 'Технологические', 'Экологические']
        if message.text not in Categories:
            msg = bot.reply_to(message, f'Неверный формат, повторите ввод категории', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_insertProjectCategory_step, projectName, projectDescription)
            return 

        projectCategory = message.text

        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        markup.add('🔴 Повторить ввод', '🟢 Все верно')
        
        user = getUserById(message.from_user.id)
        fioUser = filter(f'{user["lastname"] if user["lastname"] != None else ""} {user["firstname"] if user["firstname"] != None else ""} {user["patronymic"] if user["patronymic"] != None else ""} (@{user["username"]})')
        
        msg = bot.send_message(message.chat.id, f''' Новый проект добавлен, проверьте правильность заполнения данных:
                            
*Наименование проекта*: _{filter(projectName)}_

*Описание*: _{filter(projectDescription)}_

*Категория*: _{filter(projectCategory)}_

*Заявитель*: {fioUser}
    ''', parse_mode="MarkdownV2", reply_markup=markup)
        bot.register_next_step_handler(msg, process_isRepeatFillingProject_step, projectName, projectDescription, projectCategory)
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_isRepeatFillingProject_step(message, projectName, projectDescription, projectCategory):
    try:
        if message.text == "🔴 Повторить ввод":
            announceProject(message.chat.id)
        elif message.text == "🟢 Все верно":
            cursor.execute(f'''INSERT INTO projects (name, description, category) VALUES (%s, %s, %s) RETURNING id;''', (projectName, projectDescription, projectCategory))
            newProjectId = cursor.fetchone()[0]
            connection.commit()
            cursor.execute(f'''INSERT INTO users_projects (projectid, userid, role) VALUES (%s, %s, %s);''', (newProjectId, message.from_user.id, 'AUTHOR'))
            connection.commit()
            
            exitStepHandler(message, "ok")
        else:   
            msg = bot.send_message(chat_id=message.chat.id, text="Проверьте корректность данных", parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_isRepeatFillingProject_step, projectName, projectDescription, projectCategory)

    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")


# -------------------------------------------------- Create event steps -----------------------

def process_meetingDate_step(message):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
        
        meetingDate = message.text
        
        now = datetime.now()
        try:
            dt = datetime.strptime(meetingDate, "%d.%m.%y %H:%M")
            if dt <= now:
                msg = bot.reply_to(message, 'Неверный формат, повторите ввод даты и времени мероприятия:\n(Пример: _18.11.23 17:00_)', parse_mode="Markdown")
                bot.register_next_step_handler(msg, process_meetingDate_step)                
                return
        except Exception as e:
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод даты и времени мероприятия:\n(Пример: _18.11.23 17:00_)', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_meetingDate_step)
            return
    
        msg = bot.reply_to(message, 'Информация о мероприятии и описание:', parse_mode="Markdown")
        bot.register_next_step_handler(msg, process_descriptionEvent_step, meetingDate) 
        
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_descriptionEvent_step(message, meetingDate):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
        
        description = message.text

        if len(description) < 10:
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод информации о мероприятии и описания', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_descriptionEvent_step, meetingDate)
            return

        msg = bot.reply_to(message, 'Дедлайн голосования:\n(_День.Месяц.Год Час:Мин_)', parse_mode="Markdown")
        bot.register_next_step_handler(msg, process_deadlineEvent_step, meetingDate, description) 
        
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_deadlineEvent_step(message, meetingDate, description):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
        
        deadline = message.text

        now = datetime.now()
        try:
            dt = datetime.strptime(deadline, "%d.%m.%y %H:%M")
            deadlineMin = int(dt.strftime('%M'))

            if dt <= now or dt >= datetime.strptime(meetingDate, "%d.%m.%y %H:%M") or deadlineMin % 5 != 0:
                msg = bot.reply_to(message, 'Неверный формат, повторите ввод даты и времени окончания голосования\n(Пример: _17.11.23 21:00_)', parse_mode="Markdown")
                bot.register_next_step_handler(msg, process_deadlineEvent_step, meetingDate, description)
                return
            
        except Exception as e:
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод даты и времени окончания голосования\n(Пример: _17.11.23 21:00_)', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_deadlineEvent_step, meetingDate, description)
            return

        markup = types.ReplyKeyboardMarkup(one_time_keyboard=True, resize_keyboard=True)
        markup.add('🔴 Повторить ввод', '🟢 Все верно')

        msg = bot.send_message(message.chat.id , f'''__Опрос будет выглядеть следующим образом\. Отправляем?__
                           
*{filter(meetingDate)}* состоится новое мероприятие\!
{filter(description)}

_*Дедлайн по голосованию:* {filter(deadline)}_''', parse_mode="MarkdownV2", reply_markup=markup)
        bot.register_next_step_handler(msg, process_isRepeatFillingEvent_step, meetingDate, description, deadline) 
        
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_isRepeatFillingEvent_step(message, meetingDate, description, deadline):
    try:
        if message.text == "🔴 Повторить ввод":
            createNewEvent(message.chat.id)
        elif message.text == "🟢 Все верно":

            question = meetingDate
            options = ["Приду", "Не смогу"]
            bot.send_message(chat_id=config.RESIDENT_GROUP_ID, text=f'''*{filter(meetingDate)}* состоится новое мероприятие\!
{filter(description)}

_*Дедлайн по голосованию:* {filter(deadline)}_''', parse_mode="MarkdownV2")
            pollMessage = bot.send_poll(chat_id=config.RESIDENT_GROUP_ID, question=question, options=options, is_anonymous=False)

            meetingDate = f"{meetingDate.split('.')[1]}.{meetingDate.split('.')[0]}.{meetingDate[6:]}"
            deadline = f"{deadline.split('.')[1]}.{deadline.split('.')[0]}.{deadline[6:]}"

            cursor.execute(f'''INSERT INTO events (id, description, meetingdate, isactive, polldeadline) VALUES (%s, %s, %s, %s, %s)''', 
                           (pollMessage.json['poll']['id'], description, meetingDate, True, deadline))
            connection.commit()
            
            exitStepHandler(message, "ok")
        else:   
            msg = bot.send_message(chat_id=message.chat.id, text="Проверьте данные и сделайте выбор", parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_isRepeatFillingEvent_step, meetingDate, description, deadline)

    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def process_getUserNumForSelect_step(message, projectId, partners):
    try:
        if message.text == "↩ Выйти":
            exitStepHandler(message, "ok")
            return
    
        numOfUser = message.text.strip()

        if partners.get(str(numOfUser)) is None or not numOfUser.isdigit():
            msg = bot.reply_to(message, 'Неверный формат, повторите ввод номера\n(Пример: 1)', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_getUserNumForSelect_step, projectId, partners)
            return
    
        userFioUsernameString = str(partners[str(numOfUser)])
        userId = getUserIdByUsernameAndFIO(userFioUsernameString[userFioUsernameString.find('(') + 1 : userFioUsernameString.find(')')].strip('@'), "userName")

        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        rejectBtn = types.KeyboardButton("🔴 Отмена")
        acceptBtn = types.KeyboardButton("🟢 Удалить")
        keyboard.add(rejectBtn, acceptBtn)

        msg = bot.send_message(message.chat.id, f"Удалить пользователя *«{userFioUsernameString}»* из проекта *«{getProjectById(projectId)['name']}»*?", parse_mode="Markdown", reply_markup=keyboard)
        bot.register_next_step_handler(msg, process_deletePartner_step, projectId, userId) 
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def process_deletePartner_step(message, projectId, userId):
    try:
        if message.text == "🔴 Отмена":
            exitStepHandler(message, "ok")
        elif message.text == "🟢 Удалить":
            cursor.execute(f'''DELETE FROM users_projects WHERE 
                                projectid={projectId} and userid='{userId}';''')
            connection.commit()

            bot.send_message(message.chat.id, f"Пользователь был исключен из команды проекта.", parse_mode="Markdown", reply_markup=genKeyboard(message.from_user.id))
        else:
            msg = bot.reply_to(message, 'Вы можете *отменить* удаление или *подтвердить* его.', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_deletePartner_step, projectId, userId)
        return
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

# -------------------------------------------------- Join to project group -----------------------

def process_requestToJoin_step(message, projectId, authorId):
    try:
        if message.text == "🔴 Отмена":
            exitStepHandler(message, "ok")
        elif message.text == "🔵 Отправить запрос":

            user = getUserById(message.from_user.id)
            fioUser = filter(f'{user["lastname"] if user["lastname"] != None else ""} {user["firstname"] if user["firstname"] != None else ""} {user["patronymic"] if user["patronymic"] != None else ""} (@{user["username"]})')

            keyboard = types.InlineKeyboardMarkup(row_width=2)
            rejectBtn = types.InlineKeyboardButton("🔴 Отклонить", callback_data=f'rejectRequest_{projectId}_{message.from_user.id}')
            acceptBtn = types.InlineKeyboardButton("🟢 Принять", callback_data=f'acceptRequest_{projectId}_{message.from_user.id}')
            keyboard.row(rejectBtn, acceptBtn)

            bot.send_message(authorId, f"Пользователь {fioUser} отправил запрос на вступление в команду проекта *«{filter(getProjectById(projectId)['name'])}»*", 
                            reply_markup=keyboard, 
                            parse_mode="MarkdownV2")
            
            bot.send_message(message.from_user.id, f"Ваша заявка на вступление в команду проекта *«{filter(getProjectById(projectId)['name'])}»* была отправлена\. Скоро заявитель её рассмотрит и мы уведомим Вас о результате\.", 
                            reply_markup=genKeyboard(message.from_user.id), 
                            parse_mode="MarkdownV2")
        else:
            msg = bot.reply_to(message, 'Вы можете *подать заявку* на присоединение к команде выбранного проекта или *отменить* действие.', parse_mode="Markdown")
            bot.register_next_step_handler(msg, process_requestToJoin_step, projectId, authorId)
        return
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")


# -------------------------------------------------- Auxiliary -----------------------

def exitStepHandler(message, status):
    text = "👌"
    if status == "error": 
        text = "Произошла ошибка, попробуйте позже."

    bot.send_message(chat_id=message.chat.id, text=text, parse_mode="Markdown", reply_markup=genKeyboard(message.from_user.id))
    bot.clear_step_handler_by_chat_id(chat_id=message.chat.id)
    return

def filter(text):
    text = text.replace('_', '\_').replace('*', '\*').replace('[', '\[').replace(']', '\]').replace('(', '\(').replace(')', '\)').replace('~', '\~').replace('`', '\`').replace('>', '\>').replace('#', '\#').replace('+', '\+').replace('-', '\-').replace('=', '\=').replace('|', '\|').replace('{', '\{').replace('}', '\}').replace('.', '\.').replace('!', '\!')
    return text

def extract_arg(arg):
    return arg.split()[1:]

def updateFullname(chatId, userId):
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    goHomeBtn = types.KeyboardButton(text="↩ Выйти")
    keyboard.add(goHomeBtn)

    msg = bot.send_message(chat_id=chatId, text=f"Введите ФИО через пробел\n(_Иванов Иван Иванович_)", parse_mode="Markdown", reply_markup=keyboard)
    bot.register_next_step_handler(msg, process_updateFullName_step, userId)

def announceProject(chatId):
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    goHomeBtn = types.KeyboardButton(text="↩ Выйти")
    keyboard.add(goHomeBtn)

    msg = bot.send_message(chatId, f"Введите *наименование* проекта", parse_mode="MarkdownV2", reply_markup=keyboard)
    bot.register_next_step_handler(msg, process_insertProjectName_step)

def getUsernameForUpdate(chatId):
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    goHomeBtn = types.KeyboardButton(text="↩ Выйти")
    keyboard.add(goHomeBtn)

    msg = bot.send_message(chatId, "Введите никнейм пользователя для редактирования данных\n(Пример: @volodin)", reply_markup=keyboard)        
    bot.register_next_step_handler(message=msg, callback=process_getUsernameForUpdate_step)

def getUsernameForSelect(chatId):
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    goHomeBtn = types.KeyboardButton(text="↩ Выйти")
    keyboard.add(goHomeBtn)

    msg = bot.send_message(chatId, "Введите данные пользователя в одном из форматов для вывода информации:\n(Пример: *@username*)\n(Пример: *Фамилия*)\n(Пример: *Фамилия Имя*)", reply_markup=keyboard, parse_mode="Markdown")        
    bot.register_next_step_handler(message=msg, callback=process_getUsernameForSelect_step)

def getProjectnameForSelect(chatId):
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    goHomeBtn = types.KeyboardButton(text="↩ Выйти")
    keyboard.add(goHomeBtn)

    msg = bot.send_message(chatId, "Введите наименование проекта для отображения данных\n(Пример: TimeTrace)", reply_markup=keyboard)        
    bot.register_next_step_handler(message=msg, callback=process_getProjectnameForSelect_step)

def getUserNumForSelect(chatId, projectId):
    keyboard = types.ReplyKeyboardMarkup(row_width=4, resize_keyboard=True)

    cursor.execute(f'''SELECT CONCAT(lastname, ' ', firstname, ' ', patronymic, ' (@', username, ')') FROM users_projects
                        INNER JOIN users ON users_projects.userid = users.id
                        WHERE projectid = {projectId} and role = 'PARTNER';''')
    partnersList = cursor.fetchall()

    partners = {}
    partnersStr = ""
    for i in range(0, len(partnersList)):
        partners[str(i+1)] = partnersList[i][0]
        keyboard.add(types.KeyboardButton(text=f"{i+1}"))
        partnersStr += f"{i+1}\. {filter(partnersList[i][0])}\n"

    goHomeBtn = types.KeyboardButton(text="↩ Выйти")
    keyboard.add(goHomeBtn)

    msg = bot.send_message(chatId, f"{partnersStr}\n_Введите *номер пользователя*, чтобы исключить его из проекта_ \n\(\-\-\> *1\.* Иванов Иван Иванович \(@username\)\)", parse_mode="MarkdownV2", reply_markup=keyboard)        
    bot.register_next_step_handler(msg, process_getUserNumForSelect_step, projectId, partners)

def createNewEvent(chatId):
    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)
    goHomeBtn = types.KeyboardButton(text="↩ Выйти")
    keyboard.add(goHomeBtn)

    msg = bot.send_message(chat_id=chatId, text=f"Введите дату проведения мероприятия\n(_День.Месяц.Год Час:Мин_)", parse_mode="Markdown", reply_markup=keyboard)
    bot.register_next_step_handler(msg, process_meetingDate_step)

# -------------------------------------------------- Requests -----------------------

def projectInfo(message, projectId):
    try:
        cursor.execute(f'''SELECT * FROM users_projects
                            INNER JOIN projects ON users_projects.projectid = projects.id
                            WHERE projectid = {projectId} and role = 'AUTHOR';''')
        columns = [desc[0] for desc in cursor.description]
        items = cursor.fetchall()
        author = dict(zip(columns, items[0]))

        cursor.execute(f'''SELECT username, lastname, firstname, patronymic, userid FROM users_projects
                            INNER JOIN users ON users_projects.userid = users.id
                            WHERE projectid = {projectId} and role = 'PARTNER';''')
        items = cursor.fetchall()
        partners = ""

        for partner in items:
            partners += filter(f'{partner[1] if partner[1] != None else ""} {partner[2] if partner[2] != None else ""} {partner[3] if partner[3] != None else ""} (@{partner[0]});\n')
        
        if partners.strip() != "":
            partners = f"*Партнеры*: \n{partners}"[:-2] + "\."

        user = getUserById(author['userid'])
        fioUser = filter(f'{user["lastname"] if user["lastname"] != None else ""} {user["firstname"] if user["firstname"] != None else ""} {user["patronymic"] if user["patronymic"] != None else ""} (@{user["username"]})')

        aboutProject = "__Информация о проекте:__"
        keyboard = genKeyboard(message.from_user.id)

        if partners.strip() != "" and getUserById(message.from_user.id)['status'].find("ADMIN") != -1:
            keyboard = types.InlineKeyboardMarkup(row_width=1)
            kickPartnerBtn = types.InlineKeyboardButton('Редактировать команду', callback_data=f'deletePartnerFrom_{projectId}')
            keyboard.add(kickPartnerBtn)
            aboutProject = ""
            bot.send_message(message.chat.id, "__Информация о проекте:__", parse_mode="MarkdownV2", reply_markup=genKeyboard(message.from_user.id))

        bot.send_message(message.chat.id, f'''{aboutProject}
                         
*Наименование проекта*: _{filter(author['name'])}_

*Описание*: _{filter(author['description'])}_

*Категория*: _{filter(author['category'])}_

*Заявитель*: {fioUser}

{partners}
    ''', parse_mode="MarkdownV2", reply_markup=keyboard)
    except Exception as ex:
        print("Error: ", ex)
        exitStepHandler(message, "error")

def userInfo(message, userId):
    try:
        cursor.execute(f'''SELECT name FROM users_projects
                            INNER JOIN projects ON users_projects.projectid = projects.id
                            WHERE userid = '{userId}';''')
        projects = cursor.fetchall()

        projectNames = ""
        for project in projects:
            projectNames += project[0] + ", "
        
        if projectNames.strip() != "":
            projectNames = f"*Участник проектов*\: {filter(projectNames)}"[:-2]
        user = getUserById(userId)

        additionInfo = ""
        if user['educationalinstitution'] == "РЭУ им. Г.В. Плеханова":
            additionInfo = f" ({user['course']} курс, {user['faculty']}, по направлению «{user['direction']}», специальность «{user['speciality']}» в группе {user['group']})"
        bot.send_message(message.chat.id, f'''__Информация по пользователю @{filter(user['username'])}__\:

*ФИО*: _{user['lastname'] if user['lastname'] != None else ""} {user['firstname'] if user['firstname'] != None else ""} {user['patronymic'] if user['patronymic'] != None else ""}_

*Сфера деятельности*: _{filter(user['fieldofactivity']) if user['fieldofactivity'] is not None else filter("---")}_

*Учебное заведение*: _{filter(user['educationalinstitution']) if user['educationalinstitution'] is not None else filter("---")}{filter(additionInfo) if user['educationalinstitution'] is not None else ""}_

*О себе*: _{filter(user['aboutme']) if user['aboutme'] is not None else filter("---")}_

*Телефон*: _{filter(user['phonenum']) if user['phonenum'] is not None else filter("---")}_

*Email*: {filter(user['email']) if user['email'] is not None else filter("---")}

{projectNames.strip()}''', 
        reply_markup=genKeyboard(message.from_user.id),
        parse_mode="MarkdownV2")
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

def getProjects(page):
    buttons_per_page = 8
    cursor.execute(f'''SELECT name, id FROM projects
                        ORDER BY name
                        LIMIT {buttons_per_page} OFFSET {page-1} * {buttons_per_page};''')
    items = cursor.fetchall()
    cursor.execute(f'''SELECT COUNT(*) FROM projects;''')
    countOfProjects = cursor.fetchone()[0]
    return items, countOfProjects

def getUserById(userId):
    cursor.execute(f'''SELECT * FROM users 
                        WHERE id = '{userId}';''')

    columns = [desc[0] for desc in cursor.description]
    items = cursor.fetchone()
    user = dict(zip(columns, items))
    return user

def getProjectById(projectId):
    cursor.execute(f'''SELECT * FROM projects 
                        WHERE id = '{projectId}';''')
    
    columns = [desc[0] for desc in cursor.description]
    items = cursor.fetchone()
    project = dict(zip(columns, items))
    return project

def getUserIdByUsernameAndFIO(userInfoMsg, dataFormat):
    
    match dataFormat:
        case "userName":
            cursor.execute(f'''SELECT EXISTS(SELECT 1 FROM users
                                    WHERE username = '{userInfoMsg}');''')
            exists = cursor.fetchone()[0]
            if not exists: return "-"

            cursor.execute(f'''SELECT id FROM users 
                        WHERE username = '{userInfoMsg}';''')
            userId = cursor.fetchone()[0]
            return userId 
        case "lastName":
            cursor.execute(f'''SELECT EXISTS(SELECT 1 FROM users
                                    WHERE lastname = '{userInfoMsg}');''')
            exists = cursor.fetchone()[0]
            if not exists: return "-"

            cursor.execute(f'''SELECT id FROM users 
                        WHERE lastName = '{userInfoMsg}';''')
            userId = cursor.fetchall()
            return userId 
        case "lastName_firstName":
            cursor.execute(f'''SELECT EXISTS(SELECT 1 FROM users
                        WHERE lastname = '{userInfoMsg.split(' ')[0]}' and firstname = '{userInfoMsg.split(' ')[1]}');''')
            exists = cursor.fetchone()[0]
            if not exists: return "-"

            cursor.execute(f'''SELECT id FROM users 
                        WHERE lastname = '{userInfoMsg.split(' ')[0]}' and firstname = '{userInfoMsg.split(' ')[1]}';''')
            userId = cursor.fetchone()[0]
            return userId 
    
def getProjectIdByProjectname(name):
    cursor.execute(f'''SELECT EXISTS(SELECT 1 FROM projects
                                WHERE name = '{name}');''')
    exists = cursor.fetchone()[0]
    if not exists:
        return "-"

    cursor.execute(f'''SELECT id FROM projects 
                        WHERE name = '{name}';''')
    projectId = cursor.fetchone()[0]
    return projectId 


# -------------------------------------------------- Keyboards and markup -----------------------

def create_inline_keyboard(items, page):
    countOfProjects = getProjects(1)[1]

    buttons_per_page = 8  
    start_idx = (page - 1) * buttons_per_page
    
    leftBtn = countOfProjects - start_idx
    end_idx = start_idx + leftBtn if leftBtn < buttons_per_page else start_idx + buttons_per_page

    keyboard = types.InlineKeyboardMarkup(row_width=2)

    for i in range(0, len(items), 2):
        button1 = types.InlineKeyboardButton(items[i][0], callback_data=f'project_{items[i][1]}')
        if i + 1 == len(items):
            keyboard.add(button1)
            break
        button2 = types.InlineKeyboardButton(items[i+1][0], callback_data=f'project_{items[i+1][1]}')
        keyboard.row(button1, button2)

    if page > 1 and end_idx < countOfProjects:
        prev_button = types.InlineKeyboardButton('⬅', callback_data=f'prev_{page}')
        next_button = types.InlineKeyboardButton('➡', callback_data=f'next_{page}')
        keyboard.row(prev_button, next_button)
    elif page == 1 and leftBtn <= buttons_per_page:
        announceProject_button = types.InlineKeyboardButton('✉ Заявить о новом проекте', callback_data=f'announce_project')
        keyboard.add(announceProject_button)
    elif end_idx < countOfProjects:
        next_button = types.InlineKeyboardButton('➡', callback_data=f'next_{page}')
        keyboard.add(next_button)
    elif page > 1:
        announceProject_button = types.InlineKeyboardButton('✉ Заявить о новом проекте', callback_data=f'announce_project')
        prev_button = types.InlineKeyboardButton('⬅', callback_data=f'prev_{page}')
        keyboard.add(announceProject_button)
        keyboard.add(prev_button)

    return keyboard

def genKeyboard(userId):
    user = getUserById(userId)

    keyboard = types.ReplyKeyboardMarkup(row_width=1, resize_keyboard=True)

    if user['status'].find("RESIDENT") != -1:
        editInfoBtnText = "📝 Заполнить информацию о себе"
        if user['fieldofactivity'] != None and str(user['fieldofactivity']).strip() != "":
            editInfoBtnText = "📝 Редактировать информацию о себе"

        updateProfileBtn = types.KeyboardButton(text=editInfoBtnText)
        updateProjectGroupBtn = types.KeyboardButton(text="🗂 Добавить информацию о проекте")
        keyboard.add(updateProfileBtn, updateProjectGroupBtn)
    if user['status'].find("ADMIN") != -1:
        adminPanelBtn = types.KeyboardButton(text="🛠️ Панель администратора")
        keyboard.add(adminPanelBtn)
    if user['status'].find("EVENT_MANAGER") != -1:
        newEventBtn = types.KeyboardButton(text="🎟️ Новое мероприятие")
        keyboard.add(newEventBtn)
    if user['status'].find("USER") != -1:
        updateFullNameBtn = types.KeyboardButton(text="📃 Таблица вакансий")
        keyboard.add(updateFullNameBtn)
    return keyboard


# -------------------------------------------------- Callback handlers -----------------------

@bot.callback_query_handler(func=lambda call: call.data.startswith('prev_') or call.data.startswith('next_'))
def handle_navigation(call):
    countOfProjects = getProjects(1)[1]
    action, page = call.data.split('_')

    if action == 'prev':
        current_page = max(1, int(page) - 1)
    elif action == 'next':
        current_page = min((countOfProjects - 1) // 8 + 1, int(page) + 1)

    projects = getProjects(current_page)[0]

    keyboard = create_inline_keyboard(projects, current_page)
    bot.edit_message_reply_markup(call.message.chat.id, call.message.message_id, reply_markup=keyboard)

@bot.callback_query_handler(func=lambda call: call.data.startswith('project_'))
def handle_project_selection(call):
    try:
        bot.clear_step_handler_by_chat_id(chat_id=call.message.chat.id)

        idProject = call.data[8:]

        keyboard = types.ReplyKeyboardMarkup(row_width=2, resize_keyboard=True)
        cancelReqBtn = types.KeyboardButton(text="🔴 Отмена")
        approveReqBtn = types.KeyboardButton(text="🔵 Отправить запрос")
        keyboard.row(cancelReqBtn, approveReqBtn)

        cursor.execute(f'''SELECT * FROM users_projects
                            INNER JOIN projects ON users_projects.projectid = projects.id
                            WHERE projectid = {idProject} and role = 'AUTHOR';''')

        columns = [desc[0] for desc in cursor.description]
        items = cursor.fetchall()
        author = dict(zip(columns, items[0]))

        cursor.execute(f'''SELECT username, lastname, firstname, patronymic, userid FROM users_projects
                            INNER JOIN users ON users_projects.userid = users.id
                            WHERE projectid = {idProject} and role = 'PARTNER';''')

        items = cursor.fetchall()
        partners = ""
        alreadyMember = False 
        isAuthor = False 

        if str(author['userid']) == str(call.from_user.id):
            alreadyMember = True
            isAuthor = True

        for partner in items:
            if (str(partner[4]) == str(call.from_user.id) or str(author['userid']) == str(call.from_user.id)):
                alreadyMember = True
            partners += filter(f'{partner[1] if partner[1] != None else ""} {partner[2] if partner[2] != None else ""} {partner[3] if partner[3] != None else ""} (@{partner[0]});\n')
        
        if partners.strip() != "":
            partners = f"*Партнеры*: \n{partners}"[:-2] + "\."

        user = getUserById(author['userid'])
        fioUser = filter(f'{user["lastname"] if user["lastname"] != None else ""} {user["firstname"] if user["firstname"] != None else ""} {user["patronymic"] if user["patronymic"] != None else ""} (@{user["username"]})')
        
        aboutProject = f"__Вы можете подать запрос на вступление в команду проекта «{filter(author['name'])}»\.__"
        if alreadyMember:
            aboutProject = "__Вы уже состоите в команде этого проекта\.__"
            keyboard = genKeyboard(call.from_user.id)
            if partners.strip() != "" and (isAuthor or getUserById(call.from_user.id)['status'].find("ADMIN") != -1):
                aboutProject = ""
                keyboard = types.InlineKeyboardMarkup(row_width=1)
                kickPartnerBtn = types.InlineKeyboardButton('Редактировать команду', callback_data=f'deletePartnerFrom_{idProject}')
                keyboard.add(kickPartnerBtn)
                bot.send_message(call.message.chat.id,"__Вы уже состоите в команде этого проекта\.__", parse_mode="MarkdownV2", reply_markup=genKeyboard(call.from_user.id))
        msg = bot.send_message(call.message.chat.id, f''' {aboutProject}
                            
*Наименование проекта*: _{filter(author['name'])}_

*Описание*: _{filter(author['description'])}_

*Категория*: _{filter(author['category'])}_

*Заявитель*: {fioUser}

{partners}
    ''', parse_mode="MarkdownV2", reply_markup=keyboard)
        
        if alreadyMember: return
        bot.register_next_step_handler(msg, process_requestToJoin_step, idProject, author['userid'])

    except Exception as e:
            print(e)
            exitStepHandler(call.message, "error")

@bot.callback_query_handler(func=lambda call: call.data == 'announce_project')
def handle_announceProject(call):
    try:
        announceProject(call.message.chat.id)
    except Exception as e:
        print(e)
        exitStepHandler(call.message, "error")

@bot.callback_query_handler(func=lambda call: call.data == 'update_user_info')
def handle_updateUserInfo(call):
    try:
        getUsernameForUpdate(call.message.chat.id)
    except Exception as e:
        print(e)
        exitStepHandler(call.message, "error")

@bot.callback_query_handler(func=lambda call: call.data == 'get_user_info')
def handle_getUserInfo(call):
    try:
        getUsernameForSelect(call.message.chat.id)
    except Exception as e:
        print(e)
        exitStepHandler(call.message, "error")

@bot.callback_query_handler(func=lambda call: call.data == 'get_project_info')
def handle_getProjectInfo(call):
    try:
        getProjectnameForSelect(call.message.chat.id)
    except Exception as e:
        print(e)
        exitStepHandler(call.message, "error")

@bot.callback_query_handler(func=lambda call: call.data == 'get_users_excel')
def handle_getUsersExcel(call):
    try:
        query_string = '''SELECT id, username, lastname, firstname, patronymic, 
                                fieldofactivity, aboutme, educationalinstitution, faculty, course, direction, 
                                speciality, "group", phonenum, email, status FROM users ORDER BY lastname'''
        filepath = "usersList.xlsx"

        export_to_excel(query_string,("Имя пользователя", "Фамилия", "Имя", "Отчество", 
                                        "Род деятельности", "О себе", "Образовательное учреждение", "Факультет", "Курс", "Направление", 
                                        "Специальность", "Группа", "Номер тел.", "Email", "Роли", "Участник проектов"), filepath)

        with open('usersList.xlsx', 'rb') as tmp:
            bot.send_document(call.from_user.id, tmp)
    except Exception as e:
        print(e)
        exitStepHandler(call.message, "error")

@bot.callback_query_handler(func=lambda call: call.data == 'get_projects_excel')
def handle_getProjectsExcel(call):
    try:
        query_string = '''SELECT * FROM projects ORDER BY name'''
        filepath = "projectsList.xlsx"

        export_to_excel(query_string,("Наименование", "Описание", "Категория", "Заявитель", "Партнеры"), filepath)

        with open('projectsList.xlsx', 'rb') as tmp:
            bot.send_document(call.from_user.id, tmp)
    except Exception as e:
        print(e)
        exitStepHandler(call.message, "error")

def export_to_excel(query_string, headings, filepath):
    cursor.execute(query_string)
    data = cursor.fetchall()

    wb = openpyxl.Workbook()
    sheet = wb.active

    header = NamedStyle(name="header")
    header.font = Font(bold=True)
    header.border = Border(bottom=Side(border_style="thin"))
    header.alignment = Alignment(horizontal="center", vertical="center")

    for colno, heading in enumerate(headings, start = 1):
        sheet.cell(row = 1, column = colno).value = heading

    header_row = sheet[1]
    for cell in header_row:
        cell.style = header

    for rowno, row in enumerate(data, start = 2):
        itemId = row[0]
        for colno, cell_value in enumerate(row[1:], start = 1):
            if filepath.startswith('usersList'):
                if colno == 15: # projectMember - 1 column
                    cursor.execute(f'''SELECT name FROM users_projects
                            INNER JOIN projects ON users_projects.projectid = projects.id
                            WHERE userid = '{itemId}';''')
                    projects = cursor.fetchall()
                    projectNames = ""
                    for project in projects:
                        projectNames += project[0] + ", "
                    
                    sheet.cell(row = rowno, column = colno).value = cell_value # status
                    sheet.cell(row = rowno, column = colno+1).value = projectNames[:-2] # projectsMember
                    continue
            if filepath.startswith('usersList'):
                if colno == 15: # projectMember - 1 column
                    cursor.execute(f'''SELECT name FROM users_projects
                            INNER JOIN projects ON users_projects.projectid = projects.id
                            WHERE userid = '{itemId}';''')
                    projects = cursor.fetchall()
                    projectNames = ""
                    for project in projects:
                        projectNames += project[0] + ", "
                    
                    sheet.cell(row = rowno, column = colno).value = cell_value # status
                    sheet.cell(row = rowno, column = colno+1).value = projectNames[:-2] # projectsMember
                    continue
            elif filepath.startswith('projectsList'):
                if colno == 3: # projectMember - 1 column
                    cursor.execute(f'''SELECT concat(lastname, ' ', firstname, ' ', patronymic, ' (@', username, ')') as FIO FROM users_projects
                                        INNER JOIN users ON users_projects.userid = users.id
                                        WHERE projectid = {itemId} and role = 'AUTHOR';''')
                    fioAuthor = cursor.fetchone()[0]

                    cursor.execute(f'''SELECT concat(lastname, ' ', firstname, ' ', patronymic, ' (@', username, ')') FROM users_projects
                            INNER JOIN users ON users_projects.userid = users.id
                            WHERE projectid = {itemId} and role = 'PARTNER';''')
                    items = cursor.fetchall()
                    partners = ""

                    for partner in items:
                        partners += f'{partner[0] if partner[0] != None else ""};\n'
                    
                    if partners.strip() != "":
                        partners = f"{partners}"[:-2] + "."

                    sheet.cell(row = rowno, column = colno).value = cell_value
                    sheet.cell(row = rowno, column = colno+1).value = fioAuthor
                    sheet.cell(row = rowno, column = colno+2).value = partners
                    continue
            sheet.cell(row = rowno, column = colno).value = cell_value
    wb.save(filepath)

@bot.callback_query_handler(func=lambda call: call.data.startswith('rejectRequest_') or call.data.startswith('acceptRequest_'))
def handle_acceptRejectRequest(call):
    try:
        projectId = str(call.data).split('_')[1]
        userId = str(call.data).split('_')[2]
        
        if(call.data.startswith('rejectRequest_')):
            bot.send_message(userId, f"🔴 Ваша заявка на вступленние в проект *«{getProjectById(projectId)['name']}»* была отклонена.", 
                                parse_mode="Markdown")
            bot.send_message(call.from_user.id, f"Заявка отклонена.", 
                                parse_mode="Markdown")

        elif (call.data.startswith('acceptRequest_')):
            isAlreadyMember = False
            cursor.execute(f'''SELECT EXISTS(SELECT 1 FROM users_projects
                                WHERE projectid = {projectId} and userid = '{userId}');''')
            isAlreadyMember = cursor.fetchone()

            if isAlreadyMember[0]:
                return
            
            cursor.execute(f"INSERT INTO users_projects (projectid, userid, role) VALUES (%s, %s, %s);", (projectId, userId, 'PARTNER'))
            connection.commit()

            bot.send_message(userId, f"🟢 Ваша заявка на вступленние в проект *«{getProjectById(projectId)['name']}»* была одобрена.", 
                                parse_mode="Markdown")
            
            bot.send_message(call.from_user.id, f"Заявка одобрена.", 
                                parse_mode="Markdown")
    except Exception as e:
        print(e)
        exitStepHandler(call.message, "error")

@bot.callback_query_handler(func=lambda call: call.data.startswith('need_memo_'))
def handle_setNeedMemo(call):
    try:
        userId = str(call.data).split('_')[2]
        eventId = str(call.data).split('_')[3]

        cursor.execute(f'''UPDATE events_users SET needmemo = {True} 
                       WHERE eventid = '{eventId}' and userid = '{userId}';''')
        connection.commit()
        
        bot.send_message(userId, "Запрос на служебную записку оформлен.")
    except Exception as e:
        print(e)
        exitStepHandler(call.message, "error")

@bot.callback_query_handler(func=lambda call: call.data.startswith('need_pass_'))
def handle_setNeedPass(call):
    try:
        userId = str(call.data).split('_')[2]
        eventId = str(call.data).split('_')[3]

        cursor.execute(f'''UPDATE events_users SET needpass = {True} 
                       WHERE eventid = '{eventId}' and userid = '{userId}';''')
        connection.commit()
        
        bot.send_message(userId, "Запрос на пропуск оформлен.")
    except Exception as e:
        print(e)
        exitStepHandler(call.message, "error")

@bot.callback_query_handler(func=lambda call: call.data.startswith('deletePartnerFrom_'))
def handle_deletePartnerFrom(call):
    try:
        projectId = str(call.data).split('_')[1]

        getUserNumForSelect(call.message.chat.id, projectId)
    except Exception as e:
        print(e)
        exitStepHandler(call.message, "error")


    # try:
    #     projectId = str(call.data).split('_')[1]

    #     cursor.execute(f'''UPDATE events_users SET needpass = {True} 
    #                    WHERE eventid = '{eventId}' and userid = '{userId}';''')
    #     connection.commit()
        
    #     msg = bot.send_message(call.from_user.id, "Запрос на пропуск оформлен.")
    #     bo
    # except Exception as e:
    #     print(e)
    #     exitStepHandler(call.message, "error")


# -------------------------------------------------- Commands -----------------------

@bot.message_handler(commands=['start'])
def start(message):
    
    # user_id = message.from_user.id
    # username = message.from_user.username
    # firstName = message.from_user.first_name
    # lastName = message.from_user.last_name

    # cursor.execute('''INSERT INTO users (id, username, firstName, lastName, status) VALUES (%s, %s, %s, %s, '{RESIDENT}') 
    #                         On CONFLICT(id) DO UPDATE
    #                         SET (username, firstName, lastName, status) = (EXCLUDED.username, EXCLUDED.firstName, EXCLUDED.lastName, EXCLUDED.status);''', 
    #                         (user_id, username, firstName, lastName))

    # connection.commit()

    user = getUserById(message.from_user.id)

    helloTxt = "🐝 Привет\!"

    if user['status'].find("RESIDENT") != -1:
        helloTxt = "🐝 Привет, резидент\!"
    if user['status'].find("ADMIN") != -1:
        helloTxt = '''🐝 Привет\! Быстрые команды:
*Обновить информацию о пользователе*: 
`\/updateUserInfo @username`
*Просмотреть информацию о пользователе*: 
`\/getUserInfo @username / Фамилия / Фамилия Имя`
*Просмотреть информацию о проекте*: 
`\/getProjectInfo projectName`'''
 
    bot.send_message(chat_id=message.chat.id, text= helloTxt, reply_markup=genKeyboard(message.from_user.id), parse_mode="MarkdownV2")

@bot.message_handler(commands=['updateUserInfo'])
def updateResidentInfo(message):
    try: 
        if getUserById(message.from_user.id)['status'].find("ADMIN") == -1:
            bot.send_message(chat_id=message.chat.id, text= f"Недостаточно прав")
            return 
        
        username = extract_arg(message.text)[0].strip('@')
        userId = getUserIdByUsernameAndFIO(username, "userName")
        if userId == "-":
            bot.send_message(chat_id=message.chat.id, text= f"Пользователь не найден")
        else:
            bot.send_message(chat_id=message.chat.id, text= f"Обновление данных пользователя @{username}")
            updateFullname(message.chat.id, userId)

    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

@bot.message_handler(commands=['getUserInfo'])
def getUserInfo(message):
    try:
        if getUserById(message.from_user.id)['status'].find("ADMIN") == -1:
            bot.send_message(chat_id=message.chat.id, text= f"Недостаточно прав")
            return 

        userInfoMsg = message.text[13:].strip()

        dataFormat = ""
        if userInfoMsg.startswith('@'):
            userInfoMsg = userInfoMsg.strip('@')
            dataFormat = "userName"
        elif len(userInfoMsg.split(' ')) == 1:
            dataFormat = "lastName"
        elif len(userInfoMsg.split(' ')) == 2:
            dataFormat = "lastName_firstName"
        else:
            bot.send_message(chat_id=message.chat.id, text= f"Пользователь не найден")
            return

        userId = getUserIdByUsernameAndFIO(userInfoMsg, dataFormat)
        if userId == "-":
            bot.send_message(chat_id=message.chat.id, text= f"Пользователь не найден")
            return
        
        if not isinstance(userId, str):
            for user in userId:
                userInfo(message, user[0])
        else:
            userInfo(message, userId)
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

@bot.message_handler(commands=['getProjectInfo'])
def getProjectInfo(message):
    try:
        if getUserById(message.from_user.id)['status'].find("ADMIN") == -1:
            bot.send_message(chat_id=message.chat.id, text= f"Недостаточно прав")
            return 
        
        projectName = extract_arg(message.text)[0].strip()
        projectId = getProjectIdByProjectname(projectName)
        if projectId == "-":
            bot.send_message(chat_id=message.chat.id, text= f"Проект не найден")
        projectInfo(message, projectId)
    except Exception as e:
        print(e)
        exitStepHandler(message, "error")

# -------------------------------------------------- Message handlers -----------------------

@bot.message_handler(func=lambda message: message.text == "📝 Заполнить информацию о себе" or message.text == "🔴 Повторить ввод" or message.text == "📝 Редактировать информацию о себе")
def updateFullName_handler(message):

    user = getUserById(message.from_user.id)
    if (user['status'].find("RESIDENT") == -1):
        return
    updateFullname(message.chat.id, message.from_user.id)
   
@bot.message_handler(func=lambda message: message.text == "🗂 Добавить информацию о проекте")
def selectionProjectGroup(message):

    user = getUserById(message.from_user.id)
    if (user['status'].find("RESIDENT") == -1):
        return

    page = 1
    inlineKeyboard = create_inline_keyboard(getProjects(page)[0], page)
    msg = bot.send_message(chat_id=message.chat.id, text=f"Выберите проект, в команде которого Вы состоите или добавьте свой", parse_mode="Markdown", reply_markup=inlineKeyboard)
    # bot.register_next_step_handler(message=msg, callback=process_updateProjectGroup_step)

@bot.message_handler(func=lambda message: message.text == "🟢 Все верно" or message.text == "↩ Выйти")
def goMainMenu(message):
    keyboard = genKeyboard(message.from_user.id)
    match message.text:
        case "🟢 Все верно":
            bot.send_message(chat_id=message.chat.id, text=f"👍", parse_mode="Markdown", reply_markup=keyboard)
        case "↩ Выйти":
            bot.send_message(chat_id=message.chat.id, text=f"👌", parse_mode="Markdown", reply_markup=keyboard)

@bot.message_handler(func=lambda message: message.text == "🛠️ Панель администратора")
def getAdminPanel(message):

    user = getUserById(message.from_user.id)
    if (user['status'].find("ADMIN") == -1):
        return

    keyboard = types.InlineKeyboardMarkup(row_width=1)
    updateUserInfoBtn = types.InlineKeyboardButton('✏️ Обновить информацию о пользователе', callback_data=f'update_user_info')
    getUserInfoBtn = types.InlineKeyboardButton('👤 Информация о пользователе ', callback_data=f'get_user_info')
    getProjectInfoBtn = types.InlineKeyboardButton('📗 Информация о проекте', callback_data=f'get_project_info')
    getUsersExcelBtn = types.InlineKeyboardButton('📁 Выгрузить Excel по пользователям', callback_data=f'get_users_excel')
    getProjectsExcelBtn = types.InlineKeyboardButton('🗃 Выгрузить Excel по проектам', callback_data=f'get_projects_excel')

    keyboard.add(updateUserInfoBtn, getUserInfoBtn, getProjectInfoBtn, getUsersExcelBtn, getProjectsExcelBtn)

    bot.send_message(chat_id=message.chat.id, text=f'''__Возможности администратора__''', parse_mode="MarkdownV2", reply_markup=keyboard)

@bot.message_handler(func=lambda message: message.text == "🎟️ Новое мероприятие")
def newEvent(message):

    user = getUserById(message.from_user.id)
    if (user['status'].find("EVENT_MANAGER") == -1):
        return

    createNewEvent(message.chat.id)

@bot.message_handler(content_types=["new_chat_members"])
def handler_new_member(message):

    user_id = message.new_chat_members[0].id
    username = message.new_chat_members[0].username
    firstName = message.new_chat_members[0].first_name
    lastName = message.new_chat_members[0].last_name

    try:
        if(str(message.chat.id) == '-4066956508'):
            cursor.execute('''INSERT INTO users (id, username, firstName, lastName, status) VALUES (%s, %s, %s, %s, '{RESIDENT}') 
                            On CONFLICT(id) DO UPDATE
                            SET (username, firstName, lastName, status) = (EXCLUDED.username, EXCLUDED.firstName, EXCLUDED.lastName, EXCLUDED.status);''', 
                            (user_id, username, firstName, lastName))

        connection.commit()

        bot.send_message(message.chat.id, f"Привет, {firstName}")

    except Exception as ex:
        print('[INFO] Error postgresql ', ex)

@bot.message_handler(commands=['send'])
def send_message_to_group(message):

    chat_id = message.chat.id
    # members_count = bot.get_chat_members_count(chat_id)
    print(chat_id)

    # poll_message.json
    # with open('mytest.csv', newline='', encoding='utf-8') as csvfile:
    #     reader = csv.DictReader(csvfile)
    
    #     for row in reader:
    #         user_id = row['user id']
    #         if user_id != '6370696970':
    #             try:
    #                 bot.send_message(user_id, message.text.replace('/send ', ''))
    #             except telebot.apihelper.ApiTelegramException as e:
    #                 print(f"Not delivered. User {row['username']} hasnt yet started a conversation.")

@bot.poll_answer_handler()
def handle_poll_answer(pollAnswer):
    isGoingToCome = False
    needMemo = False
    needPass = False

    match str(pollAnswer.option_ids):
        case '[0]':
            isGoingToCome = True
            markup = types.InlineKeyboardMarkup(row_width=1)
            needMemoBtn = types.InlineKeyboardButton('📄 Запросить служебную записку', callback_data=f'need_memo_{pollAnswer.user.id}_{pollAnswer.poll_id}')
            needPassBtn = types.InlineKeyboardButton('🎫 Нужен пропуск на территорию', callback_data=f'need_pass_{pollAnswer.user.id}_{pollAnswer.poll_id}')

            user = getUserById(pollAnswer.user.id)
            if user['educationalinstitution'] == 'РЭУ им. Г.В. Плеханова':
                markup.add(needMemoBtn)
            else:
                markup.add(needPassBtn)

            bot.send_message(chat_id=pollAnswer.user.id, text=f"Не забудь про мероприятие {getMeetingDateEventById(pollAnswer.poll_id).strftime('%d.%m.%Y %H:%M')}.", parse_mode="Markdown", reply_markup=markup)

        case '[1]':
            isGoingToCome = False
        case '[]':
            isGoingToCome = False

    cursor.execute(f'''INSERT INTO events_users (eventid, userid, isgoingtocome, needmemo, needpass) VALUES (%s, %s, %s,%s, %s) 
                        ON CONFLICT (eventid, userid)
                        DO UPDATE SET (isgoingtocome, needmemo, needpass) = (EXCLUDED.isgoingtocome, EXCLUDED.needmemo, EXCLUDED.needpass);''', 
                    (pollAnswer.poll_id, pollAnswer.user.id, isGoingToCome, needMemo, needPass))
    connection.commit()


def getMeetingDateEventById(eventId):
    cursor.execute(f"SELECT meetingdate FROM events WHERE id='{eventId}'")
    meetingDate = cursor.fetchone()[0]

    return meetingDate

def schedule_checker():
    while True:
        schedule.run_pending()
        time.sleep(1)

def isTimeToNewsletterForDocManager():
    now = datetime.now() 
    current_time = now.strftime("%Y-%m-%d %H:%M")
    # print(current_time)
    print(current_time)
    cursor.execute(f'''SELECT id, polldeadline FROM events WHERE 
                        EXISTS (SELECT polldeadline FROM events 
        	            WHERE isactive=true and polldeadline='{current_time}');''')
    docManagerIds = cursor.fetchall()

    if len(docManagerIds) > 0:
        for id in docManagerIds:
            
            cursor.execute("SELECT id FROM users WHERE status && '{DOCUMENT_MANAGER}';")
            docManagerId = cursor.fetchone()[0]

            cursor.execute(f'''SELECT COUNT(*) FROM events_users WHERE eventid='{id[0]}' and isgoingtocome=true;''')
            countOfPlanningToCome = cursor.fetchone()[0]
            cursor.execute(f'''SELECT COUNT(*) FROM events_users 
                           INNER JOIN users ON events_users.userid=users.id
                           WHERE eventid='{id[0]}' and isgoingtocome=true and educationalinstitution='РЭУ им. Г.В. Плеханова';''')
            countOfFromREA = cursor.fetchone()[0]
            countOfFromAnother = countOfPlanningToCome - countOfFromREA
            cursor.execute(f'''SELECT COUNT(*) FROM events_users WHERE eventid='{id[0]}' and isgoingtocome=true and needmemo=true;''')
            countOfneedMemo = cursor.fetchone()[0]
            cursor.execute(f'''SELECT COUNT(*) FROM events_users WHERE eventid='{id[0]}' and isgoingtocome=true and needpass=true;''')
            countOfneedPass = cursor.fetchone()[0]

            bot.send_message(docManagerId, f'''На мероприятие _{filter(str(getMeetingDateEventById(id[0]).strftime('%d.%m.%Y %H:%M')))}_:
*Зарегистрировалось:* {countOfPlanningToCome} чел\.
*Резиденты из РЭУ:* {countOfFromREA} чел\.
*Резиденты из других ВУЗов:* {countOfFromAnother} чел\.
*Запросили служебную записку:* {countOfneedMemo} чел\.
*Запросили пропуск на территорию:* {countOfneedPass} чел\.''', parse_mode="MarkdownV2")


if __name__ == '__main__':
    schedule.every().hour.at(":10").do(isTimeToNewsletterForDocManager)
    schedule.every().hour.at(":15").do(isTimeToNewsletterForDocManager)

    schedule.every().hour.at(":20").do(isTimeToNewsletterForDocManager)
    schedule.every().hour.at(":25").do(isTimeToNewsletterForDocManager)

    schedule.every().hour.at(":30").do(isTimeToNewsletterForDocManager)
    schedule.every().hour.at(":35").do(isTimeToNewsletterForDocManager)

    schedule.every().hour.at(":40").do(isTimeToNewsletterForDocManager)
    schedule.every().hour.at(":45").do(isTimeToNewsletterForDocManager)

    schedule.every().hour.at(":50").do(isTimeToNewsletterForDocManager)
    schedule.every().hour.at(":55").do(isTimeToNewsletterForDocManager)

    schedule.every().hour.at(":00").do(isTimeToNewsletterForDocManager)
    schedule.every().hour.at(":05").do(isTimeToNewsletterForDocManager)

    Thread(target=schedule_checker).start()
    bot.infinity_polling()

# bot.infinity_polling()



